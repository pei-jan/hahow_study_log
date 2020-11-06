#!/usr/bin/env python
# coding: utf-8

# In[13]:


def hahow學習紀錄():
    print('建課程')
    print()
    import os
    import csv
    path='/Users/pei/Documents/PCX/hahow學習紀錄/課程學員上課狀況/'

    documents=os.listdir(path)
    print(documents)

    import openpyxl
    import xlrd
    import time
    wb = openpyxl.load_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/格式檔/建課程格式.xlsx')
    ws = wb.active
    class_list = []
    class_time_list = {}
    for i in documents:

        if i[0] != '.':
            path2='/Users/pei/Documents/PCX/hahow學習紀錄/課程學員上課狀況/'+i
            files=os.listdir(path2)
            for j in files:
                name_list = j.split('_')
                class_name = '[HAHOW學習紀錄]' + name_list[0]
                class_list.append(class_name)
                with open(path2+'/'+j,newline = '') as csvfile:
                    rows = csv.reader(csvfile)

                    class_time = [row[3] for row in rows]
                    class_time_list[class_name] = class_time[1]
    class_list_only = []
    for k in class_list:
        if not k in class_list_only:
            class_list_only.append(k)


    #原課程代碼表
    wb2 = openpyxl.load_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/課程代碼清單.xlsx')
    ws2 = wb2.active
    classid ={}
    classbook = xlrd.open_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/課程代碼清單.xlsx')
    sheet = classbook.sheet_by_index(0)
    for i in range(len(sheet.col_values(0))):
        classid[ws2['A'+str(i+1)].value] = ws2['B'+str(i+1)].value

    cell = 0
    class_code = 0
    class_code_input = input('輸入課程代碼末一碼')
    for m in range(len(class_list_only)):
        if class_list_only[m] in classid:
            continue
        ws['A'+ str(cell+3) ] = class_list_only[m]
        #if class_list_only[m] in classid:
        #ws['B'+ str(m+3) ] = classid[class_list_only[m]]
        #else:    
        ws['B'+ str(cell+3) ] = '#hahow'+str(class_code)+class_code_input
        ws['C'+ str(cell+3) ] ='HAHOW課程紀錄'
        ws['H'+ str(cell+3) ] ='HAHOW課程紀錄'
        ws['I'+ str(cell+3) ] ='這是你在HAHOW平台上的學習紀錄'
        ws['L'+ str(cell+3) ] ='HAHOW課程紀錄'
        ws['K'+ str(cell+3) ] ='#COE00'
        ws['P'+ str(cell+3) ] ='HAHOW'
        ws['V'+ str(cell+3) ] = class_time_list[class_list_only[m]].replace(' 分鐘','')
        class_code += 1
        cell += 1
    print(str(cell)+'堂課')

    wb.save('/Users/pei/Documents/PCX/hahow學習紀錄/python/建課程上傳.xlsx')
    if cell > 0:
        print()
        print('調整課程代碼')
        print()

        class_list3 =[]
        wb3 = openpyxl.load_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/建課程上傳.xlsx')
        ws3 = wb3.active
        classbook3 = xlrd.open_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/建課程上傳.xlsx')
        sheet3 = classbook3.sheet_by_index(0)
        for i in range(2 , len(sheet3.col_values(0))):
            name_list3 = ws3['A'+str(i+1)].value.split(']')
            class_name3 = name_list3[1]
            class_list3.append(class_name3)


        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        import sys
        import os
        import time
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.common.action_chains import ActionChains
        import pyperclip
        import pyautogui

        browser = webdriver.Chrome()
        browser.get('https://business.hahow.in/home')
        browser.maximize_window()
        browser.find_element_by_xpath("//*[@id='app']/main/div/div[1]/div/div[1]/form/div[1]/div[2]/div/span/input").send_keys('peijan796@cathaylife.com.tw')
        browser.find_element_by_xpath("//*[@id='app']/main/div/div[1]/div/div[1]/form/div[2]/div[2]/div/span/input").send_keys('Anc799633')
        browser.find_element_by_xpath("//*[@id='app']/main/div/div[1]/div/div[1]/form/div[4]/div/div/span/div/button").click()

        for i in range(len(class_list3)):
            browser.switch_to.window(browser.window_handles[0])
            browser.find_element_by_xpath("//*[@id='app']/header/nav/div/ul[1]/li/form/input[1]").send_keys(class_list3[i])
            time.sleep(3)
            pyautogui.press('enter')
            time.sleep(2)
            browser.find_element_by_xpath("//*[@id='app']/header/nav/div/ul[1]/li/form/input[1]").clear()
            browser.find_element_by_xpath("//*[@id='app']/header/nav/div/ul[1]/li/form/input[1]").send_keys(class_list3[i])
            time.sleep(3)
            pyautogui.press('enter')
            time.sleep(5)
            x = len(browser.find_elements_by_partial_link_text(class_list3[i]))
            for p in range(x):
                search1 = browser.find_elements_by_partial_link_text(class_list3[i])[p].text
                if search1[0:len(class_list3[i])] == class_list3[i]:
                    browser.find_elements_by_partial_link_text(class_list3[i])[p].click()
            time.sleep(5)
            url1 = browser.current_url
            print(url1)
            class_code3 = url1.split('/')
            new_class_name3 = class_code3[-1]
            ws['B'+str(i+3)] = '#hahow'+str(new_class_name3).zfill(3)
            wb.save('/Users/pei/Documents/PCX/hahow學習紀錄/python/建課程上傳update.xlsx')


        os.remove('/Users/pei/Documents/PCX/hahow學習紀錄/python/建課程上傳.xlsx')

        print()
        print('建課程代碼表')
        print()

        import os
        import openpyxl
        import xlrd
        import time

        #建課程代號對照表
        classdic ={}
        classbook = xlrd.open_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/建課程上傳update.xlsx')
        sheet = classbook.sheet_by_index(0)
        for i in range(2 , len(sheet.col_values(0))):
            cell_value_class = sheet.cell(i,0).value
            cell_value_id = sheet.cell(i,1).value
            classdic[cell_value_class] = cell_value_id

        classdic.update(classid)
        wb = openpyxl.load_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/格式檔/課程代碼清單空白.xlsx')
        ws = wb.active

        for i in range(len(classdic)):
            ws['A'+str(i+1)] = list(classdic.keys())[i]
            ws['B'+str(i+1)] = list(classdic.values())[i]
        wb.save('/Users/pei/Documents/PCX/hahow學習紀錄/python/課程代碼清單.xlsx')

        print()
        print('建班次')
        print()

        import openpyxl
        import xlrd
        import time
        import datetime
        wb = openpyxl.load_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/格式檔/建班次格式.xlsx')
        ws = wb.active

        wb2 = openpyxl.load_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/建課程上傳update.xlsx')
        ws2 = wb2.active


        classbook = xlrd.open_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/建課程上傳update.xlsx')
        sheet = classbook.sheet_by_index(0)

        x=datetime.datetime.today()   
        startdate = str(x.year) + '/' +str(x.month) +'/' +str(x.day-2)  #發佈開始日兩天前

        for i in range(len(sheet.col_values(0))-2):
            ws['A'+str(i+3)] = ws2['B'+str(i+3)].value
            ws['C'+str(i+3)] = ws2['B'+str(i+3)].value
            ws['B'+str(i+3)] = '紀錄'
            ws['D'+str(i+3)] = 'COE00'
            ws['G'+str(i+3)] = startdate
            ws['Z'+str(i+3)] = '0'  #是否發佈
            ws['AA'+str(i+3)] = '0'  #是否結案
            ws['AB'+str(i+3)] = '0'  #新開/補登
        wb.save('/Users/pei/Documents/PCX/hahow學習紀錄/python/建班次上傳.xlsx')

    print()
    print('建學員')
    print()

    classdic ={}
    classbook = xlrd.open_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/課程代碼清單.xlsx')
    sheet = classbook.sheet_by_index(0)
    wb2 = openpyxl.load_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/課程代碼清單.xlsx')
    ws2 = wb2.active
    for i in range(len(sheet.col_values(0))):
        classdic[ws2['A'+str(i+1)].value] = ws2['B'+str(i+1)].value

    import csv

    path='/Users/pei/Documents/PCX/hahow學習紀錄/課程學員上課狀況/'
    documents=os.listdir(path)
    wb = openpyxl.load_workbook('/Users/pei/Documents/PCX/hahow學習紀錄/python/格式檔/建學員格式.xlsx')
    ws = wb.active

    study_time = []
    student_name = []
    student_email= []
    ifpass = []
    class_id = []
    company = []
    for i in documents:
        if i[0] != '.':
            path2='/Users/pei/Documents/PCX/hahow學習紀錄/課程學員上課狀況/'+i
            files=os.listdir(path2)
            for j in files:
                with open(path2+'/'+j,newline = '') as csvfile:
                    rows = csv.reader(csvfile)
                    for row in rows:
                        if rows.line_num == 1 or row[5] == '' or row[11].upper() == 'FALSE':
                            continue
                        taketime = ''
                        time = int(taketime.join([x for x in row[5] if x.isdigit()]))*60
                        study_time.append(time)
                        student_name.append(row[0])
                        student_email.append(row[1]) 
                        passtrue =row[11].upper().replace('TRUE','2')
                        ifpass.append(passtrue)
                        name_list = j.split('_')
                        class_name = '[HAHOW學習紀錄]' + name_list[0]
                        class_id.append(classdic[class_name])
                        company.append(i)
    for i in range(len(student_name)):
        ws['A'+str(i+3)] = class_id[i]
        ws['B'+str(i+3)] = student_name[i]
        ws['C'+str(i+3)] = '0'
        ws['E'+str(i+3)] = ifpass[i]
        ws['F'+str(i+3)] = study_time[i]
        ws['H'+str(i+3)] = student_email[i]
        ws['I'+str(i+3)] = company[i]
    wb.save('/Users/pei/Documents/PCX/hahow學習紀錄/python/建學員上傳.xlsx')


# In[ ]:




